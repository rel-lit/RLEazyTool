"""
Excel 处理模块 - 负责将游戏数据保存到Excel文件（图片嵌入模式）
"""
import os
import logging
from openpyxl import Workbook, load_workbook
from openpyxl import styles
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
import io
import requests

from utils import get_excel_path, is_file_open, logger

logger = logging.getLogger(__name__)


class ExcelHandler:
    """Excel文件处理器"""
    
    def __init__(self):
        self.filepath = get_excel_path()
        self.workbook = None
        self.sheet = None
    
    def _create_new_workbook(self):
        """创建新的工作簿并初始化表头"""
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = 'Steam游戏数据'
        
        # 设置表头 - 使用图片嵌入模式
        headers = [
            '封面图片', '游戏名', '价格', '好评率', 
            '标签', '标签',
            '商店链接', '语言', 
            '评分员1', '短评', '分数',
            '评分员2', '短评', '分数',
            '评分员3', '短评', '分数',
            '评分员4', '短评', '分数',
            '平均分'
        ]
        self.sheet.append(headers)
        
        # 设置列宽 - 为图片列设置合适的宽度
        self.sheet.column_dimensions['A'].width = 30  # 封面图片（列宽30）
        self.sheet.column_dimensions['B'].width = 25  # 游戏名
        self.sheet.column_dimensions['C'].width = 10  # 价格
        self.sheet.column_dimensions['D'].width = 10  # 好评率
        self.sheet.column_dimensions['E'].width = 10  # 标签1
        self.sheet.column_dimensions['F'].width = 10  # 标签2
        self.sheet.column_dimensions['G'].width = 60  # 商店链接
        self.sheet.column_dimensions['H'].width = 10  # 语言
        for col in ['I', 'L', 'O', 'R']:
            self.sheet.column_dimensions[col].width = 12
        for col in ['J', 'M', 'P', 'S']:
            self.sheet.column_dimensions[col].width = 25
        for col in ['K', 'N', 'Q', 'T']:
            self.sheet.column_dimensions[col].width = 8
        self.sheet.column_dimensions['U'].width = 10
        
        # 设置表头样式
        for cell in self.sheet[1]:
            cell.font = cell.font.copy(bold=True)
            cell.alignment = styles.Alignment(horizontal='center', vertical='center')
        
        logger.info("创建新的Excel工作簿")
    
    def _load_existing_workbook(self):
        """加载已存在的工作簿"""
        try:
            self.workbook = load_workbook(self.filepath)
            self.sheet = self.workbook.active
            logger.info("加载现有Excel文件")
        except Exception as e:
            logger.warning(f"加载Excel文件失败，将创建新文件：{str(e)}")
            self._create_new_workbook()
    
    def _check_file_availability(self):
        """检查文件是否可用（未被其他程序打开）"""
        if os.path.exists(self.filepath):
            if is_file_open(self.filepath):
                logger.error(f"Excel文件已被打开，请先关闭文件：{self.filepath}")
                return False
        return True
    
    def _download_image_from_url(self, image_url):
        """从URL下载图片并返回PIL Image对象"""
        try:
            if not image_url:
                return None
            
            logger.info(f"正在下载图片: {image_url}")
            response = requests.get(image_url, timeout=10)
            response.raise_for_status()
            
            # 将字节数据转换为PIL Image对象
            img = PILImage.open(io.BytesIO(response.content))
            logger.info("图片下载成功")
            return img
        except Exception as e:
            logger.error(f"图片下载失败: {str(e)}")
            return None
    
    def _resize_image_to_fit_cell(self, pil_image, target_width=220, target_height=150):
        """调整图片尺寸以完全贴合单元格（列宽30，行高40）"""
        if pil_image is None:
            return None
        
        # 直接调整图片到目标尺寸，完全贴合单元格
        # 列宽30 ≈ 220像素，行高40 ≈ 150像素
        resized_img = pil_image.resize((target_width, target_height), PILImage.LANCZOS)
        
        return resized_img
    
    def _add_game_row(self, game_data):
        """添加一行游戏数据，包含嵌入式图片"""
        tags = game_data.get('tags', [])
        tag1 = tags[0] if len(tags) > 0 else ''
        tag2 = tags[1] if len(tags) > 1 else ''
        
        row_data = [
            '',  # A: 封面图片（稍后插入图片对象）
            game_data.get('name', '未知'),  # B: 游戏名
            game_data.get('price', '未知'),  # C: 价格
            game_data.get('review', '暂无评测'),  # D: 好评率
            tag1,  # E: 标签1
            tag2,  # F: 标签2
            game_data.get('url', ''),  # G: 商店链接
            game_data.get('languages', '未知'),  # H: 语言
            '', '', '',  # I, J, K: 评分员1
            '', '', '',  # L, M, N: 评分员2
            '', '', '',  # O, P, Q: 评分员3
            '', '', '',  # R, S, T: 评分员4
            ''  # U: 平均分
        ]
        
        self.sheet.append(row_data)
        row_num = self.sheet.max_row
        
        # 设置该行所有单元格垂直居中
        for col_idx in range(1, len(row_data) + 1):
            cell = self.sheet.cell(row=row_num, column=col_idx)
            cell.alignment = styles.Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # 尝试下载并插入图片
        cover_image_url = game_data.get('cover_image')
        if cover_image_url:
            pil_image = self._download_image_from_url(cover_image_url)
            if pil_image:
                # 调整图片大小以适应单元格
                resized_img = self._resize_image_to_fit_cell(pil_image)
                if resized_img:
                    # 将PIL Image转换为openpyxl兼容的Image对象
                    img_stream = io.BytesIO()
                    resized_img.save(img_stream, format='PNG')
                    img_stream.seek(0)
                    
                    openpyxl_img = Image(img_stream)
                    # 设置图片在单元格中的位置
                    cell_ref = f'A{row_num}'
                    openpyxl_img.anchor = cell_ref
                    
                    # 设置固定行高40
                    self.sheet.row_dimensions[row_num].height = 40  # 固定行高40
                    
                    # 添加图片到工作表
                    self.sheet.add_image(openpyxl_img, cell_ref)
                    logger.info(f"在第{row_num}行插入图片")
        
        logger.info(f"添加游戏数据到第{row_num}行：{game_data.get('name')}")
        return row_num
    
    def save_game_data(self, game_data):
        """保存游戏数据到Excel（图片嵌入模式）"""
        if not self._check_file_availability():
            return False
        
        if os.path.exists(self.filepath):
            self._load_existing_workbook()
        else:
            self._create_new_workbook()
        
        try:
            row_num = self._add_game_row(game_data)
            self.workbook.save(self.filepath)
            logger.info(f"数据保存成功：{self.filepath}")
            return True
        except Exception as e:
            logger.error(f"保存Excel文件失败：{str(e)}")
            return False
    
    def get_current_row_count(self):
        """获取当前数据行数（不包括表头）"""
        if os.path.exists(self.filepath):
            try:
                wb = load_workbook(self.filepath)
                ws = wb.active
                return max(0, ws.max_row - 1)
            except:
                return 0
        return 0
